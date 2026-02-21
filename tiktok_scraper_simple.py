"""
TikTok Trending Hashtags Scraper
================================
Fetches trending hashtags from TikTok's Creative Center.

Strategy (in priority order):
  1. Hit TikTok Creative Center's internal API directly (fast, structured JSON)
  2. Scrape the server-rendered HTML from the Creative Center page (no browser needed)
  3. Fall back to a curated static database (always works)

Designed to run in GitHub Actions with zero browser dependencies.
Only requires: requests, beautifulsoup4, pandas, openpyxl
"""

import json
import re
import time
import requests
from datetime import datetime
from bs4 import BeautifulSoup

# ─────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────

COUNTRY_CODE = ""          # Empty = all regions (matches the default Creative Center view)
PERIOD = 7                 # 7-day trending window
MAX_HASHTAGS = 100         # How many to try to fetch

# TikTok Creative Center industry filter codes (from the site's dropdown)
# These are the numeric IDs TikTok uses internally for industry filtering.
INDUSTRY_IDS = {
    "All":                    "",
    "Apparel & Accessories":  "2",
    "Beauty & Personal Care": "3",
    "Education":              "6",
    "Financial Services":     "9",
    "Food & Beverage":        "15",
    "Games":                  "18",
    "Health":                 "36",
    "Home Improvement":       "21",
    "News & Entertainment":   "24",
    "Pets":                   "28",
    "Sports & Outdoor":       "31",
    "Tech & Electronics":     "33",
    "Travel":                 "35",
    "Vehicle & Transportation": "37",
}

# Map TikTok industry names → your website's JSON keys
WEBSITE_KEY_MAP = {
    "All":                    "general",
    "Beauty & Personal Care": "lifestyle",
    "Food & Beverage":        "food",
    "Health":                 "fitness",
    "Apparel & Accessories":  "fashion",
    "Tech & Electronics":     "tech",
    "Games":                  "tech",        # Merge gaming into tech
    "Travel":                 "travel",
    "Financial Services":     "business",
    "News & Entertainment":   "entertainment",
    "Sports & Outdoor":       "sports",
    "Education":              "education",
    "Home Improvement":       "home",
    "Pets":                   "pets",
    "Vehicle & Transportation": "auto",
}

# Standard headers to look like a normal browser
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://ads.tiktok.com/business/creativecenter/inspiration/popular/hashtag/pc/en",
}


# ─────────────────────────────────────────────
# Method 1: TikTok Creative Center Internal API
# ─────────────────────────────────────────────

def fetch_via_api(industry_name="All", industry_id="", country_code="", period=7, limit=50):
    """
    Call TikTok Creative Center's internal API endpoint.
    This is the same endpoint the website's JavaScript calls.
    Returns a list of hashtag strings, or None on failure.
    """
    url = "https://ads.tiktok.com/creative_radar_api/v1/popular_trend/hashtag/list"

    params = {
        "period":       period,
        "country_code": country_code,
        "page":         1,
        "limit":        limit,
        "sort_by":      "popular",
    }
    if industry_id:
        params["industry_id"] = industry_id

    try:
        resp = requests.get(url, headers=HEADERS, params=params, timeout=15)
        if resp.status_code != 200:
            print(f"  [API] HTTP {resp.status_code} for {industry_name}")
            return None

        data = resp.json()

        # The API returns: { "code": 0, "data": { "list": [ { "hashtag_name": "...", ... }, ... ] } }
        if data.get("code") != 0:
            print(f"  [API] Non-zero response code: {data.get('code')} — {data.get('msg', '')}")
            return None

        items = data.get("data", {}).get("list", [])
        if not items:
            print(f"  [API] Empty list for {industry_name}")
            return None

        hashtags = []
        for item in items:
            name = item.get("hashtag_name", "").strip()
            if name:
                tag = f"#{name}" if not name.startswith("#") else name
                hashtags.append(tag)

        print(f"  [API] ✅ Got {len(hashtags)} hashtags for {industry_name}")
        return hashtags

    except Exception as e:
        print(f"  [API] Error for {industry_name}: {e}")
        return None


# ─────────────────────────────────────────────
# Method 2: Scrape the server-rendered HTML
# ─────────────────────────────────────────────

def fetch_via_html(country_code="", period=7):
    """
    Fetch the Creative Center hashtags page and parse the SSR HTML.
    The page renders ~20 hashtags in the initial HTML without JavaScript.
    Returns a list of hashtag strings, or None on failure.
    """
    url = "https://ads.tiktok.com/business/creativecenter/inspiration/popular/hashtag/pc/en"

    params = {}
    if country_code:
        params["countryCode"] = country_code
    if period:
        params["period"] = period

    try:
        resp = requests.get(url, headers=HEADERS, params=params, timeout=20)
        if resp.status_code != 200:
            print(f"  [HTML] HTTP {resp.status_code}")
            return None

        soup = BeautifulSoup(resp.text, "html.parser")
        hashtags = []

        # Strategy A: Look for hashtag text in the structured list items
        # The page renders hashtags as "# hashtagname" in heading/link elements
        for el in soup.find_all(["a", "div", "span", "h3", "p"]):
            text = el.get_text(strip=True)
            # Match patterns like "# shabebarat" or "#home"
            if text.startswith("# ") or text.startswith("#"):
                tag = text.replace("# ", "#").strip()
                if tag.startswith("#") and len(tag) > 1 and " " not in tag:
                    if tag.lower() not in [h.lower() for h in hashtags]:
                        hashtags.append(tag)

        # Strategy B: Parse from links like /hashtag/HASHTAGNAME/pc/en
        if len(hashtags) < 5:
            for link in soup.find_all("a", href=True):
                href = link["href"]
                match = re.search(r"/hashtag/([^/?]+)", href)
                if match:
                    name = match.group(1)
                    tag = f"#{name}"
                    if tag.lower() not in [h.lower() for h in hashtags]:
                        hashtags.append(tag)

        # Strategy C: Look for JSON data embedded in script tags
        if len(hashtags) < 5:
            for script in soup.find_all("script"):
                script_text = script.string or ""
                # Look for hashtag_name in JSON blobs
                for match in re.finditer(r'"hashtag_name"\s*:\s*"([^"]+)"', script_text):
                    tag = f"#{match.group(1)}"
                    if tag.lower() not in [h.lower() for h in hashtags]:
                        hashtags.append(tag)

        if hashtags:
            print(f"  [HTML] ✅ Parsed {len(hashtags)} hashtags from page HTML")
            return hashtags
        else:
            print("  [HTML] No hashtags found in page HTML")
            return None

    except Exception as e:
        print(f"  [HTML] Error: {e}")
        return None


# ─────────────────────────────────────────────
# Method 3: Static fallback database
# ─────────────────────────────────────────────

def get_fallback_data():
    """Last resort: hardcoded hashtag sets that are always valid."""
    return {
        "general": [
            "#fyp", "#foryou", "#viral", "#trending", "#foryoupage",
            "#tiktok", "#xyzbca", "#tiktokviral", "#viralvideo", "#fypシ",
        ],
        "fitness": [
            "#fitness", "#workout", "#gym", "#fitfam", "#healthylifestyle",
            "#weightloss", "#fitnessmotivation", "#health", "#wellness",
            "#fittok", "#gymtok", "#motivation",
        ],
        "food": [
            "#food", "#foodie", "#cooking", "#recipe", "#foodtok",
            "#baking", "#chef", "#foodporn", "#homecooking", "#easyrecipe",
            "#cookingtiktok", "#yummy",
        ],
        "lifestyle": [
            "#beauty", "#makeup", "#skincare", "#beautytips", "#makeuptutorial",
            "#beautytok", "#skincareroutine", "#grwm", "#beautyhacks",
            "#skintok", "#makeupartist",
        ],
        "fashion": [
            "#fashion", "#style", "#ootd", "#outfit", "#streetstyle",
            "#fashiontok", "#fashioninspo", "#outfitinspo", "#styleinspo",
            "#fashiontrends", "#fashionblogger",
        ],
        "tech": [
            "#tech", "#technology", "#gaming", "#gamer", "#ai",
            "#techtok", "#gamedev", "#pc", "#gamingsetup", "#techreview",
            "#esports", "#twitch",
        ],
        "travel": [
            "#travel", "#traveltok", "#adventure", "#wanderlust", "#vacation",
            "#explore", "#traveling", "#travelgram", "#travelphotography",
            "#travelblogger", "#travelvlog",
        ],
        "business": [
            "#business", "#entrepreneur", "#marketing", "#money", "#investing",
            "#financetok", "#businesstips", "#hustle", "#stocks",
            "#entrepreneurship", "#sidehustle", "#moneytok",
        ],
    }


# ─────────────────────────────────────────────
# Main orchestrator
# ─────────────────────────────────────────────

def scrape_all_hashtags():
    """
    Try each method in priority order. Build a complete hashtag dataset
    organized by website category keys.
    Returns (dict, source_label) where dict maps category→list of hashtag strings.
    """
    website_data = {}
    source = "unknown"

    # ── Attempt 1: API method (try each industry) ──
    print("\n🔌 Attempting TikTok Creative Center API...")
    api_success_count = 0

    for industry_name, industry_id in INDUSTRY_IDS.items():
        web_key = WEBSITE_KEY_MAP.get(industry_name)
        if not web_key:
            continue

        tags = fetch_via_api(
            industry_name=industry_name,
            industry_id=industry_id,
            country_code=COUNTRY_CODE,
            period=PERIOD,
            limit=50,
        )
        if tags:
            # Merge into existing key (e.g., Games + Tech both map to "tech")
            if web_key in website_data:
                existing = set(t.lower() for t in website_data[web_key])
                for t in tags:
                    if t.lower() not in existing:
                        website_data[web_key].append(t)
                        existing.add(t.lower())
            else:
                website_data[web_key] = tags
            api_success_count += 1

        # Be polite — small delay between API calls
        time.sleep(0.5)

    if api_success_count > 0:
        source = "api"
        print(f"\n✅ API method succeeded for {api_success_count}/{len(INDUSTRY_IDS)} industries")

    # ── Attempt 2: HTML scrape (general trending only) ──
    if "general" not in website_data or len(website_data.get("general", [])) < 5:
        print("\n🌐 Attempting HTML scrape of Creative Center page...")
        html_tags = fetch_via_html(country_code=COUNTRY_CODE, period=PERIOD)
        if html_tags:
            website_data["general"] = html_tags
            if source == "unknown":
                source = "html"
            else:
                source = "api+html"
            print(f"✅ HTML method got {len(html_tags)} general trending hashtags")

    # ── Attempt 3: Static fallback ──
    if not website_data or all(len(v) < 3 for v in website_data.values()):
        print("\n⚠️  Live methods failed. Using static fallback data.")
        website_data = get_fallback_data()
        source = "fallback"
    else:
        # Fill in any missing categories from fallback so the site always has data
        fallback = get_fallback_data()
        for key, tags in fallback.items():
            if key not in website_data or len(website_data[key]) < 3:
                print(f"  Filling missing category '{key}' from fallback")
                website_data[key] = tags

    return website_data, source


def save_to_json(website_data, source, filename="hashtags.json"):
    """Save the hashtag data as JSON for the website to consume."""
    output = {
        "_meta": {
            "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "source": source,
            "country": COUNTRY_CODE or "all",
            "period_days": PERIOD,
        }
    }
    # Add each category
    for key, tags in website_data.items():
        output[key] = tags

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    total = sum(len(v) for k, v in website_data.items() if k != "_meta")
    print(f"\n📄 Saved {filename}")
    print(f"   Categories: {len(website_data)}")
    print(f"   Total hashtags: {total}")
    print(f"   Source: {source}")


def save_to_excel(website_data, source, filename=None):
    """Save to formatted Excel file (optional — nice for manual review)."""
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("⚠️  pandas/openpyxl not installed — skipping Excel export")
        return

    if filename is None:
        filename = f"tiktok_hashtags_{datetime.now().strftime('%Y%m%d')}.xlsx"

    rows = []
    for category, tags in website_data.items():
        for rank, tag in enumerate(tags, 1):
            rows.append({
                "rank": rank,
                "hashtag": tag,
                "category": category,
                "scraped_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source": source,
            })

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Hashtags", index=False)

        wb = writer.book
        ws = writer.sheets["All Hashtags"]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

        # Per-category sheets
        for category in df["category"].unique():
            cat_df = df[df["category"] == category]
            sheet_name = category[:31]  # Excel 31-char limit
            cat_df.to_excel(writer, sheet_name=sheet_name, index=False)

            cat_ws = writer.sheets[sheet_name]
            for cell in cat_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    print(f"📊 Saved {filename} ({len(df)} rows across {df['category'].nunique()} categories)")


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  TikTok Trending Hashtags Scraper")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Scrape
    website_data, source = scrape_all_hashtags()

    # Save JSON (this is what the website reads)
    save_to_json(website_data, source, "hashtags.json")

    # Save Excel (nice for review / archives)
    save_to_excel(website_data, source)

    print(f"\n{'=' * 60}")
    print(f"  ✅ Done! Source: {source}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
